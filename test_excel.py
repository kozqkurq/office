from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# シートを追加
ws2 = wb.create_sheet("increment")
for y in range(0, 10):
        ws2.append([x+y*10 for x in range(1, 11)])

# Save the file
wb.save("sample.xlsx")